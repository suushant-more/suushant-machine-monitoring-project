import threading
import socket
import requests
import json
import sqlite3
import time
import datetime
import pytz
import pandas as pd
import openpyxl

from bokeh.models import HoverTool

from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO, StringIO
from flask import send_file

from flask import Flask, render_template, request, jsonify, redirect, url_for
from bokeh.models import DatetimeTicker, DatetimeTickFormatter
from datetime import datetime, timezone, timedelta
from bokeh.plotting import figure
from bokeh.embed import components
from bokeh.models import ColumnDataSource, HoverTool
from bokeh.layouts import column
from bokeh.transform import jitter
from scipy.ndimage import gaussian_filter1d
from bokeh.models import Range1d

# Set timezone
IST = pytz.timezone('Asia/Kolkata')

app = Flask(__name__)
latest_data = {}
data_lock = threading.Lock()

def create_connection():
    return sqlite3.connect('sensor_data.db', check_same_thread=False)

def initialize_database():
    conn = create_connection()
    with conn:
        c = conn.cursor()
        c.execute('''
            CREATE TABLE IF NOT EXISTS sensor_data (
                timestamp TEXT DEFAULT CURRENT_TIMESTAMP,
                machine_id TEXT,
                machine_name TEXT,
                temperature REAL,
                humidity REAL,
                current_r REAL,
                current_y REAL,
                current_b REAL,
                power_factor REAL
            )
        ''')
        conn.commit()

initialize_database()

def handle_client(conn, addr):
    print(f"Connected by {addr}")
    try:
        data = conn.recv(1024)
        if data:
            print(f"Received: {data.decode()}")
            store_data(data.decode())

            payload = json.loads(data.decode())
            with data_lock:
                global latest_data
                latest_data = {
                    'machine_id': payload['machine_id'],
                    'machine_name': payload['machine_name'],
                    'temperature': payload['temperature'],
                    'humidity': payload['humidity'],
                    'current_r': payload['current_r'],
                    'current_y': payload['current_y'],
                    'current_b': payload['current_b'],
                    'power_factor': payload['power_factor']
                }
            print(f"Latest data updated: {latest_data}")
    finally:
        conn.close()

def start_server():
    host = '0.0.0.0'
    port = 65437

    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.bind((host, port))
        s.listen()
        print(f"Listening on {host}:{port}")

        while True:
            conn, addr = s.accept()
            client_thread = threading.Thread(target=handle_client, args=(conn, addr))
            client_thread.start()

def store_data(data):
    try:
        payload = json.loads(data)
        conn = create_connection()
        with conn:
            c = conn.cursor()
            current_time = datetime.now(IST).strftime('%Y-%m-%d %H:%M:%S')
            c.execute('''
                INSERT INTO sensor_data (timestamp, machine_id, machine_name, temperature, humidity, current_r, current_y, current_b, power_factor)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (current_time, payload['machine_id'], payload['machine_name'], payload['temperature'], payload['humidity'], payload['current_r'], payload['current_y'], payload['current_b'], payload['power_factor']))
            conn.commit()
        print("Data stored in database successfully")
    except sqlite3.Error as e:
        print(f"SQLite error: {e}")
    except json.JSONDecodeError as e:
        print(f"JSON decode error: {e}")
def generate_weekly_report(plant_id=None):
    conn = create_connection()
    with conn:
        c = conn.cursor()
        last_week = (datetime.now(IST) - timedelta(weeks=1)).strftime('%Y-%m-%d')
        query = '''
            SELECT * FROM sensor_data 
            WHERE timestamp >= ? {condition}
            ORDER BY timestamp
        '''
        condition = ""
        if plant_id:
            condition = "AND machine_name LIKE ?"
            c.execute(query.format(condition=condition), (last_week, f'{plant_id}%'))
        else:
            c.execute(query.format(condition=condition), (last_week,))
        
        rows = c.fetchall()

    if rows:
        df = pd.DataFrame(rows, columns=['Timestamp', 'Machine ID', 'Machine Name', 'Temperature', 'Humidity', 'Current R', 'Current Y', 'Current B', 'Power Factor'])
        return df
    else:
        return None

def get_latest_data_from_db(machine_id):
    conn = create_connection()
    with conn:
        c = conn.cursor()
        # Query to get the most recent data available for the given machine_id
        query = '''
            SELECT timestamp, temperature, humidity, current_r, current_y, current_b, machine_name 
            FROM sensor_data 
            WHERE machine_id=? 
            ORDER BY timestamp DESC 
            LIMIT 100  -- Fetch the most recent 100 records or adjust as needed
        '''
        c.execute(query, (machine_id,))
        rows = c.fetchall()
        
        if rows:
            timestamps = [row[0] for row in rows]
            temperatures = [row[1] for row in rows]
            humidities = [row[2] for row in rows]
            current_r = [row[3] for row in rows]
            current_y = [row[4] for row in rows]
            current_b = [row[5] for row in rows]
            machine_name = rows[0][6] if rows else "Unknown"
            return timestamps, temperatures, humidities, current_r, current_y, current_b, machine_name
        else:
            return [], [], [], [], [], [], "Unknown"


    
@app.route('/download_report', methods=['GET'])
def download_report():
    report_type = request.args.get('type', 'csv')  # default to 'csv' if not specified
    plant_id = request.args.get('plant_id')

    df = generate_weekly_report(plant_id)
    if df is None:
        return "No data available for the last week.", 404

    output = BytesIO()
    
    if report_type == 'excel':
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Weekly Report')
        output.seek(0)

        # Load the workbook from the BytesIO object
        wb = openpyxl.load_workbook(output)
        ws = wb.active

        # Define the fill for highlighting
        highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=8):
            temp_cell, hum_cell, current_r_cell, current_y_cell, current_b_cell = row
            if temp_cell.value and temp_cell.value > 45:
                temp_cell.fill = highlight_fill
            if hum_cell.value and hum_cell.value > 75:
                hum_cell.fill = highlight_fill
            if current_r_cell.value and current_r_cell.value > 6:
                current_r_cell.fill = highlight_fill
            if current_y_cell.value and current_y_cell.value > 6:
                current_y_cell.fill = highlight_fill
            if current_b_cell.value and current_b_cell.value > 6:
                current_b_cell.fill = highlight_fill

        # Save the changes back to the BytesIO object
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output, download_name='weekly_report.xlsx', as_attachment=True)

    else:  # default to CSV
        df.to_csv(output, index=False)
        output.seek(0)
        return send_file(output, download_name='weekly_report.csv', as_attachment=True)




@app.route('/')
def home():
    return render_template('home.html')

@app.route('/k2')
def k2_plant():
    return redirect(url_for('dashboard', plant_id='K2'))

@app.route('/k3')
def k3_plant():
    return redirect(url_for('dashboard', plant_id='K3'))

@app.route('/k4')
def k4_plant():
    return redirect(url_for('dashboard', plant_id='K4'))

@app.route('/dashboard/<plant_id>', methods=['GET', 'POST'])
def dashboard(plant_id):
    conn = create_connection()
    with conn:
        c = conn.cursor()
        c.execute('SELECT DISTINCT machine_id, machine_name FROM sensor_data WHERE machine_name LIKE ?', (f'{plant_id}%',))
        machines = c.fetchall()

    if request.method == 'POST':
        selected_machine_id = request.form.get('machine_id')
        if selected_machine_id:
            # Pass both plant_id and machine_id to the redirect
            return redirect(url_for('machine_dashboard', plant_id=plant_id, machine_id=selected_machine_id))

    return render_template('dashboard.html', machines=machines, plant_id=plant_id)
@app.route('/dashboard/<plant_id>/<machine_id>')
def machine_dashboard(plant_id, machine_id):
    timestamps, temperatures, humidities, current_r, current_y, current_b, machine_name = get_latest_data_from_db(machine_id)
    local_tz = pytz.timezone('Asia/Kolkata')
    timestamps = [local_tz.localize(datetime.strptime(t, '%Y-%m-%d %H:%M:%S')) for t in timestamps if t is not None]
    
    # Create a figure for temperature
    p = figure(title=f"Temperature Over Time for {machine_name}", x_axis_type="datetime", x_axis_label='Time', y_axis_label='Temperature (°C)', width=800, height=300)
    p.line(x=timestamps, y=temperatures, legend_label="Temperature", line_width=2)
    
    # Add hover tool for temperature
    hover = HoverTool()
    hover.tooltips = [
        ("Date", "@x{%F %T}"),  # Display the date in tooltip
        ("Temperature", "@y{0.0} °C"),  # Display temperature
    ]
    hover.formatters = {
        '@x': 'datetime',  # Use datetime format for x-axis values
    }
    p.add_tools(hover)
    
    # Add temperature threshold line at 45°C
    temperature_threshold = 45
    p.line(x=[min(timestamps), max(timestamps)], y=[temperature_threshold, temperature_threshold], color='red', line_width=2, legend_label="Threshold 45°C", line_dash="dashed")

    # Create a figure for humidity
    q = figure(title=f"Humidity Over Time for {machine_name}", x_axis_type="datetime", x_axis_label='Time', y_axis_label='Humidity (%)', width=800, height=300)
    q.line(x=timestamps, y=humidities, legend_label="Humidity", line_width=2)
    
    # Add hover tool for humidity
    hover_q = HoverTool()
    hover_q.tooltips = [
        ("Date", "@x{%F %T}"),
        ("Humidity", "@y{0.0} %"),
    ]
    hover_q.formatters = {
        '@x': 'datetime',
    }
    q.add_tools(hover_q)

    # Add humidity threshold line at 75%
    humidity_threshold = 75
    q.line(x=[min(timestamps), max(timestamps)], y=[humidity_threshold, humidity_threshold], color='brown', line_width=2, legend_label="Threshold 75%", line_dash="dashed")

    # Plotting current data with circles (R, Y, B) and adding threshold lines for current R, Y, and B
    r = figure(title=f"Current Over Time for {machine_name}", x_axis_type="datetime", x_axis_label='Time', y_axis_label='Current (A)', width=800, height=300)
    r.circle(x=timestamps, y=current_r, size=10, color="red", legend_label="Current R")
    r.square(x=timestamps, y=current_y, size=10, color="green", legend_label="Current Y")
    r.triangle(x=timestamps, y=current_b, size=10, color="blue", legend_label="Current B")
    
    # Add hover tool for current
    hover_r = HoverTool()
    hover_r.tooltips = [
        ("Date", "@x{%F %T}"),
        ("Current R", "@y{0.00} A"),
        ("Current Y", "@y{0.00} A"),
        ("Current B", "@y{0.00} A"),
    ]
    hover_r.formatters = {
        '@x': 'datetime',
    }
    r.add_tools(hover_r)

    # Add threshold lines for current R, Y, B at 6A
    current_threshold = 6
    r.line(x=[min(timestamps), max(timestamps)], y=[current_threshold, current_threshold], color='black', line_width=2, legend_label="Threshold 6A", line_dash="dashed")

    # Set x-axis to show a larger range if possible
    if timestamps:
        start_time = min(timestamps) - timedelta(hours=1)  # Adjust to see some data before the earliest timestamp
        end_time = max(timestamps) + timedelta(hours=1)    # Adjust to see some data after the latest timestamp
        p.x_range.start = start_time
        p.x_range.end = end_time

        q.x_range.start = start_time
        q.x_range.end = end_time

        r.x_range.start = start_time
        r.x_range.end = end_time

    script, div = components(column(p, q, r))

    print(f"Rendering Dashboard with data for {machine_id}")
    return render_template("machine_dashboard.html", script=script, div=div, machine_id=machine_id, machine_name=machine_name, plant_id=plant_id)


@app.after_request
def add_header(response):
    response.cache_control.no_store = True
    return response

if __name__ == "__main__":
    threading.Thread(target=start_server).start()
    app.run(host='0.0.0.0', port=5000, debug=True)
